from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import hashlib
from jose import JWTError, jwt
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from tempfile import NamedTemporaryFile


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security setup
security = HTTPBearer()
SECRET_KEY = "your-secret-key-here-change-in-production-very-long-secret-key"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    hashed_password: str
    role: str = "USER"  # "USER" or "ADMIN"
    status: str = "PENDING"  # "PENDING", "APPROVED", "REJECTED"
    approved_by: Optional[str] = None
    approved_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    username: str
    password: str

class UserLogin(BaseModel):
    username: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user_id: str
    username: str
    role: str = "USER"
    status: Optional[str] = "APPROVED"

class Pendencia(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    site: str
    ami: Optional[str] = None  # Campo AMI
    data_hora: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    tipo: str  # "Energia" or "Arcon"
    subtipo: str  # Específico baseado no tipo
    observacoes: str
    foto_base64: Optional[str] = None  # Opcional para compatibilidade com pendências legadas
    status: str = "Pendente"  # "Pendente", "Finalizado", "Validado", "Rejeitado"
    usuario_criacao: str
    usuario_finalizacao: Optional[str] = None
    data_finalizacao: Optional[datetime] = None
    informacoes_fechamento: Optional[str] = None
    foto_fechamento_base64: Optional[str] = None
    validation_status: Optional[str] = None  # "APPROVED", "REJECTED"
    validated_by: Optional[str] = None
    validated_at: Optional[datetime] = None
    validation_notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PendenciaCreate(BaseModel):
    site: str
    ami: Optional[str] = None  # Campo AMI opcional
    tipo: str
    subtipo: str
    observacoes: str
    foto_base64: str  # Obrigatória para novas pendências

class PendenciaUpdate(BaseModel):
    status: str
    informacoes_fechamento: Optional[str] = None
    foto_fechamento_base64: Optional[str] = None  # Será obrigatória na validação

class UserApproval(BaseModel):
    status: str  # "APPROVED" or "REJECTED"

class PendenciaValidation(BaseModel):
    status: str  # "APPROVED" or "REJECTED" 
    validation_notes: Optional[str] = None

class PasswordReset(BaseModel):
    new_password: str

class PasswordChange(BaseModel):
    current_password: str
    new_password: str

class FormConfig(BaseModel):
    energia_options: List[str]
    arcon_options: List[str]

class PendenciaEdit(BaseModel):
    site: str
    ami: Optional[str] = None  # Campo AMI opcional
    tipo: str
    subtipo: str
    observacoes: str
    foto_base64: Optional[str] = None

class FormConfigUpdate(BaseModel):
    energia_options: List[str]
    arcon_options: List[str]
    foto_base64: Optional[str] = None


# Auth helpers
def verify_password(plain_password, hashed_password):
    return get_password_hash(plain_password) == hashed_password

def get_password_hash(password):
    return hashlib.sha256((password + SECRET_KEY).encode()).hexdigest()

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
    
    user = await db.users.find_one({"username": username})
    if user is None:
        raise credentials_exception
    
    # Handle legacy users
    user_status = user.get("status", "APPROVED")
    user_role = user.get("role", "ADMIN" if user["username"] == "admin" else "USER")
    
    if user_status != "APPROVED":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User account not approved"
        )
    
    # Update legacy users
    if "status" not in user or "role" not in user:
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": {
                "status": user_status,
                "role": user_role
            }}
        )
        user["status"] = user_status
        user["role"] = user_role
    
    return User(**user)

async def get_admin_user(current_user: User = Depends(get_current_user)):
    if current_user.role != "ADMIN":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )
    return current_user

# Função para converter UTC para horário de Brasília
def to_brasilia_time(utc_dt: datetime) -> datetime:
    try:
        from zoneinfo import ZoneInfo
        return utc_dt.replace(tzinfo=ZoneInfo('UTC')).astimezone(ZoneInfo('America/Sao_Paulo'))
    except ImportError:
        # Fallback para sistemas sem zoneinfo
        from datetime import timedelta
        return utc_dt - timedelta(hours=3)  # UTC-3 para horário de Brasília


# Routes
@api_router.post("/register")
async def register(user_data: UserCreate):
    # Check if user already exists
    existing_user = await db.users.find_one({"username": user_data.username})
    if existing_user:
        raise HTTPException(
            status_code=400,
            detail="Username already registered"
        )
    
    # Check if this is the first user (make them admin)
    user_count = await db.users.count_documents({})
    
    # Create new user
    hashed_password = get_password_hash(user_data.password)
    user = User(
        username=user_data.username,
        hashed_password=hashed_password,
        role="ADMIN" if user_count == 0 else "USER",
        status="APPROVED" if user_count == 0 else "PENDING"
    )
    
    await db.users.insert_one(user.dict())
    
    # Create access token even for pending users (they need to see pending screen)
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    
    return Token(
        access_token=access_token,
        token_type="bearer",
        user_id=user.id,
        username=user.username,
        role=user.role,
        status=user.status
    )

@api_router.post("/login", response_model=Token)
async def login(user_data: UserLogin):
    user = await db.users.find_one({"username": user_data.username})
    if not user or not verify_password(user_data.password, user["hashed_password"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Handle legacy users without status field
    user_status = user.get("status", "APPROVED")  # Default to APPROVED for existing users
    user_role = user.get("role", "ADMIN" if user["username"] == "admin" else "USER")  # Make admin user admin
    
    if user_status == "PENDING":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Account pending admin approval"
        )
    
    if user_status == "REJECTED":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Account access denied"
        )
    
    # Update legacy users
    if "status" not in user or "role" not in user:
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": {
                "status": user_status,
                "role": user_role
            }}
        )
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user["username"]}, expires_delta=access_token_expires
    )
    
    return Token(
        access_token=access_token,
        token_type="bearer",
        user_id=user["id"],
        username=user["username"],
        role=user_role,
        status=user_status
    )

@api_router.get("/me", response_model=User)
async def read_users_me(current_user: User = Depends(get_current_user)):
    return current_user

@api_router.post("/pendencias", response_model=Pendencia)
async def create_pendencia(pendencia_data: PendenciaCreate, current_user: User = Depends(get_current_user)):
    # Validar se foto é obrigatória
    if not pendencia_data.foto_base64 or not pendencia_data.foto_base64.strip():
        raise HTTPException(status_code=400, detail="Foto é obrigatória para criar uma pendência")
    
    pendencia = Pendencia(
        site=pendencia_data.site,
        tipo=pendencia_data.tipo,
        subtipo=pendencia_data.subtipo,
        observacoes=pendencia_data.observacoes,
        foto_base64=pendencia_data.foto_base64,
        usuario_criacao=current_user.username,
        data_hora=datetime.now(timezone.utc)
    )
    
    await db.pendencias.insert_one(pendencia.dict())
    return pendencia

@api_router.get("/pendencias", response_model=List[Pendencia])
async def get_pendencias(
    site: Optional[str] = None,
    tipo: Optional[str] = None,
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if site:
        query["site"] = site
    if tipo:
        query["tipo"] = tipo
    if status:
        query["status"] = status
    
    pendencias = await db.pendencias.find(query).sort("created_at", -1).to_list(1000)
    return [Pendencia(**pendencia) for pendencia in pendencias]

@api_router.get("/sites")
async def get_sites(current_user: User = Depends(get_current_user)):
    sites = await db.pendencias.distinct("site")
    return {"sites": sites}

@api_router.put("/pendencias/{pendencia_id}", response_model=Pendencia)
async def update_pendencia(
    pendencia_id: str,
    pendencia_update: PendenciaUpdate,
    current_user: User = Depends(get_current_user)
):
    pendencia = await db.pendencias.find_one({"id": pendencia_id})
    if not pendencia:
        raise HTTPException(status_code=404, detail="Pendência not found")
    
    update_data = pendencia_update.dict(exclude_unset=True)
    if pendencia_update.status == "Finalizado":
        update_data["usuario_finalizacao"] = current_user.username
        update_data["data_finalizacao"] = datetime.now(timezone.utc)
        
        # Validar se informações de fechamento foram fornecidas
        if not pendencia_update.informacoes_fechamento or not pendencia_update.informacoes_fechamento.strip():
            raise HTTPException(status_code=400, detail="Informações de fechamento são obrigatórias")
        
        # Validar se foto de fechamento é obrigatória
        if not pendencia_update.foto_fechamento_base64 or not pendencia_update.foto_fechamento_base64.strip():
            raise HTTPException(status_code=400, detail="Foto de fechamento é obrigatória")
    
    await db.pendencias.update_one(
        {"id": pendencia_id},
        {"$set": update_data}
    )
    
    updated_pendencia = await db.pendencias.find_one({"id": pendencia_id})
    return Pendencia(**updated_pendencia)

@api_router.put("/pendencias/{pendencia_id}/edit", response_model=Pendencia)
async def edit_pendencia(
    pendencia_id: str,
    pendencia_edit: PendenciaEdit,
    current_user: User = Depends(get_current_user)
):
    pendencia = await db.pendencias.find_one({"id": pendencia_id})
    if not pendencia:
        raise HTTPException(status_code=404, detail="Pendência não encontrada")
    
    # Verificar se a pendência ainda está pendente
    if pendencia["status"] != "Pendente":
        raise HTTPException(status_code=400, detail="Só é possível editar pendências com status 'Pendente'")
    
    # Verificar se o usuário é o criador da pendência ou se é admin
    if pendencia["usuario_criacao"] != current_user.username:
        # Por simplicidade, vou permitir que qualquer usuário edite qualquer pendência pendente
        # Em produção, você poderia implementar roles de usuário
        pass
    
    update_data = pendencia_edit.dict()
    
    await db.pendencias.update_one(
        {"id": pendencia_id},
        {"$set": update_data}
    )
    
    updated_pendencia = await db.pendencias.find_one({"id": pendencia_id})
    return Pendencia(**updated_pendencia)

@api_router.delete("/pendencias/{pendencia_id}")
async def delete_pendencia(
    pendencia_id: str,
    current_user: User = Depends(get_current_user)
):
    pendencia = await db.pendencias.find_one({"id": pendencia_id})
    if not pendencia:
        raise HTTPException(status_code=404, detail="Pendência não encontrada")
    
    # Verificar se a pendência ainda está pendente
    if pendencia["status"] != "Pendente":
        raise HTTPException(status_code=400, detail="Só é possível excluir pendências com status 'Pendente'")
    
    # Verificar se o usuário é o criador da pendência ou se é admin
    if pendencia["usuario_criacao"] != current_user.username:
        # Por simplicidade, vou permitir que qualquer usuário exclua qualquer pendência pendente
        # Em produção, você poderia implementar roles de usuário
        pass
    
    result = await db.pendencias.delete_one({"id": pendencia_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pendência não encontrada")
    
    return {"message": "Pendência excluída com sucesso"}

# Admin endpoints
@api_router.get("/admin/pending-users")
async def get_pending_users(admin_user: User = Depends(get_admin_user)):
    users = await db.users.find({"status": "PENDING"}).to_list(1000)
    return [{"id": user["id"], "username": user["username"], "created_at": user["created_at"]} for user in users]

@api_router.get("/admin/all-users")
async def get_all_users(admin_user: User = Depends(get_admin_user)):
    users = await db.users.find().to_list(1000)
    return [{
        "id": user["id"], 
        "username": user["username"], 
        "role": user.get("role", "USER"),
        "status": user.get("status", "APPROVED"),
        "created_at": user["created_at"],
        "approved_by": user.get("approved_by"),
        "approved_at": user.get("approved_at")
    } for user in users]

@api_router.delete("/admin/delete-user/{user_id}")
async def delete_user(user_id: str, admin_user: User = Depends(get_admin_user)):
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Não permitir que admin exclua a si mesmo
    if user["username"] == admin_user.username:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    # Excluir usuário
    result = await db.users.delete_one({"id": user_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

@api_router.put("/admin/reset-password/{user_id}")
async def reset_password(user_id: str, password_reset: PasswordReset, admin_user: User = Depends(get_admin_user)):
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Validar nova senha
    if len(password_reset.new_password) < 4:
        raise HTTPException(status_code=400, detail="Password must be at least 4 characters")
    
    # Atualizar senha
    hashed_password = get_password_hash(password_reset.new_password)
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"hashed_password": hashed_password}}
    )
    
    return {"message": "Password reset successfully"}

@api_router.put("/admin/approve-user/{user_id}")
async def approve_user(user_id: str, approval: UserApproval, admin_user: User = Depends(get_admin_user)):
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if approval.status not in ["APPROVED", "REJECTED"]:
        raise HTTPException(status_code=400, detail="Status must be APPROVED or REJECTED")
    
    update_data = {
        "status": approval.status,
        "approved_by": admin_user.username,
        "approved_at": datetime.now(timezone.utc)
    }
    
    await db.users.update_one({"id": user_id}, {"$set": update_data})
    return {"message": f"User {approval.status.lower()} successfully"}

@api_router.get("/admin/pendencias")
async def get_all_pendencias_admin(admin_user: User = Depends(get_admin_user)):
    pendencias = await db.pendencias.find().sort("created_at", -1).to_list(1000)
    return [Pendencia(**pendencia) for pendencia in pendencias]

@api_router.put("/admin/validate-pendencia/{pendencia_id}")
async def validate_pendencia(
    pendencia_id: str,
    validation: PendenciaValidation,
    admin_user: User = Depends(get_admin_user)
):
    pendencia = await db.pendencias.find_one({"id": pendencia_id})
    if not pendencia:
        raise HTTPException(status_code=404, detail="Pendência não encontrada")
    
    update_data = {
        "validation_status": validation.status,
        "validated_by": admin_user.username,
        "validated_at": datetime.now(timezone.utc),
        "validation_notes": validation.validation_notes
    }
    
    # Se rejeitado, volta para Pendente
    if validation.status == "REJECTED":
        update_data["status"] = "Pendente"
    
    await db.pendencias.update_one({"id": pendencia_id}, {"$set": update_data})
    return {"message": "Pendência validada com sucesso"}

@api_router.delete("/admin/delete-pendencia/{pendencia_id}")
async def delete_pendencia(pendencia_id: str, admin_user: User = Depends(get_admin_user)):
    pendencia = await db.pendencias.find_one({"id": pendencia_id})
    if not pendencia:
        raise HTTPException(status_code=404, detail="Pendência não encontrada")
    
    # Admin pode excluir qualquer pendência
    result = await db.pendencias.delete_one({"id": pendencia_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pendência não encontrada")
    
    return {"message": "Pendência excluída com sucesso"}

# Endpoints para configuração do formulário
@api_router.get("/admin/form-config")
async def get_form_config(admin_user: User = Depends(get_admin_user)):
    config = await db.form_config.find_one({"type": "main"})
    if not config:
        # Configuração padrão
        default_config = {
            "type": "main",
            "energia_options": [
                "Controladora", "QDCA", "QM", "Retificador", "Disjuntor", 
                "Bateria", "Iluminação Pátio", "Sensor de Porta", 
                "Sensor de Incêndio", "Iluminação Gabinete/Container", 
                "Cabo de Alimentação"
            ],
            "arcon_options": [
                "Trocador de Calor", "Sanrio", "Walmont", "Limpeza", 
                "Contatora", "Compressor", "Gás", "Fusível", 
                "Placa Queimada", "Transformador", "Relé Térmico", 
                "Relé Falta de Fase", "Comando", "Alarme"
            ]
        }
        await db.form_config.insert_one(default_config)
        return {
            "energia_options": default_config["energia_options"],
            "arcon_options": default_config["arcon_options"]
        }
    
    return {
        "energia_options": config["energia_options"],
        "arcon_options": config["arcon_options"]
    }

@api_router.put("/admin/form-config")
async def update_form_config(config: FormConfigUpdate, admin_user: User = Depends(get_admin_user)):
    await db.form_config.update_one(
        {"type": "main"},
        {"$set": {
            "energia_options": config.energia_options,
            "arcon_options": config.arcon_options,
            "updated_by": admin_user.username,
            "updated_at": datetime.now(timezone.utc)
        }},
        upsert=True
    )
    return {"message": "Configuração do formulário atualizada com sucesso"}

# Endpoints para análise de arquivos KML
@api_router.post("/admin/upload-kml")
async def upload_kml_file(
    file: UploadFile = File(...),
    admin_user: User = Depends(get_admin_user)
):
    import xml.etree.ElementTree as ET
    import re
    
    # Validate file extension
    if not file.filename.lower().endswith('.kml'):
        raise HTTPException(status_code=400, detail="Apenas arquivos KML são aceitos")
    
    try:
        # Read file content
        content = await file.read()
        
        # Try different encodings
        try:
            content_str = content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                content_str = content.decode('utf-8-sig')  # BOM
            except UnicodeDecodeError:
                content_str = content.decode('latin1')
        
        # Clean content - remove BOM and extra whitespace
        content_str = content_str.strip()
        if content_str.startswith('\ufeff'):
            content_str = content_str[1:]
        
        # Parse KML with multiple namespace handling
        try:
            root = ET.fromstring(content_str)
        except ET.ParseError as e:
            # Try to fix common XML issues
            content_str = content_str.replace('&', '&amp;')
            root = ET.fromstring(content_str)
        
        # Get the default namespace from root
        namespace = ""
        if root.tag.startswith('{'):
            namespace = root.tag[1:root.tag.find('}')]
        
        # Define possible namespaces
        namespaces = {
            '': namespace if namespace else 'http://www.opengis.net/kml/2.2',
            'kml': 'http://www.opengis.net/kml/2.2'
        }
        
        locations = []
        
        # Try multiple ways to find placemarks
        placemarks = []
        
        # Method 1: With namespace
        for ns_prefix, ns_uri in namespaces.items():
            if ns_prefix:
                placemarks.extend(root.findall(f'.//{{{ns_uri}}}Placemark'))
            else:
                # No namespace
                placemarks.extend(root.findall('.//Placemark'))
        
        # Method 2: Search without namespace if nothing found
        if not placemarks:
            for elem in root.iter():
                if elem.tag.endswith('Placemark') or elem.tag == 'Placemark':
                    placemarks.append(elem)
        
        for placemark in placemarks:
            location_data = {}
            
            # Extract name - try multiple methods
            name = None
            for elem in placemark.iter():
                if elem.tag.endswith('name') or elem.tag == 'name':
                    if elem.text and elem.text.strip():
                        name = elem.text.strip()
                        break
            
            location_data['name'] = name or 'Unnamed Location'
            
            # Extract description
            description = None
            for elem in placemark.iter():
                if elem.tag.endswith('description') or elem.tag == 'description':
                    if elem.text and elem.text.strip():
                        description = elem.text.strip()
                        break
            
            location_data['description'] = description or ''
            
            # Extract extended data
            extended_data = {}
            for elem in placemark.iter():
                if elem.tag.endswith('ExtendedData') or elem.tag == 'ExtendedData':
                    # Look for SimpleData elements
                    for data_elem in elem.iter():
                        if data_elem.tag.endswith('SimpleData') or data_elem.tag == 'SimpleData':
                            key = data_elem.get('name', 'unknown')
                            value = data_elem.text or ''
                            extended_data[key] = value
                        elif data_elem.tag.endswith('Data') or data_elem.tag == 'Data':
                            key = data_elem.get('name', 'unknown')
                            # Look for value element inside Data
                            for value_elem in data_elem.iter():
                                if value_elem.tag.endswith('value') or value_elem.tag == 'value':
                                    extended_data[key] = value_elem.text or ''
                                    break
            
            # Add extended data to description if available
            if extended_data:
                extra_info = []
                for key, value in extended_data.items():
                    if value:
                        extra_info.append(f"{key}: {value}")
                if extra_info:
                    if location_data['description']:
                        location_data['description'] += "\n" + "\n".join(extra_info)
                    else:
                        location_data['description'] = "\n".join(extra_info)
            
            # Extract coordinates - comprehensive search
            coordinates = None
            coord_elements = []
            
            # Search for all coordinate elements
            for elem in placemark.iter():
                if elem.tag.endswith('coordinates') or elem.tag == 'coordinates':
                    if elem.text and elem.text.strip():
                        coord_elements.append(elem.text.strip())
            
            # Use first valid coordinates found
            for coord_text in coord_elements:
                coordinates = coord_text
                break
            
            if coordinates:
                # Parse coordinates (longitude,latitude,altitude format in KML)
                coord_pairs = []
                
                # Clean coordinate string
                coordinates = re.sub(r'\s+', ' ', coordinates.strip())
                
                # Split by various delimiters
                coordinate_parts = re.split(r'[\s\n\r\t]+', coordinates)
                
                for part in coordinate_parts:
                    part = part.strip()
                    if not part:
                        continue
                    
                    # Match coordinate patterns
                    # Format: longitude,latitude[,altitude]
                    coord_match = re.match(r'^(-?\d+\.?\d*),(-?\d+\.?\d*)(?:,(-?\d+\.?\d*))?$', part)
                    if coord_match:
                        try:
                            lng = float(coord_match.group(1))
                            lat = float(coord_match.group(2))
                            
                            # Validate coordinates
                            if -180 <= lng <= 180 and -90 <= lat <= 90:
                                coord_pairs.append({'lat': lat, 'lng': lng})
                        except (ValueError, TypeError):
                            continue
                
                if coord_pairs:
                    # For multiple coordinates, take the first or center
                    if len(coord_pairs) == 1:
                        location_data['latitude'] = coord_pairs[0]['lat']
                        location_data['longitude'] = coord_pairs[0]['lng']
                    else:
                        # Calculate center for multiple points
                        avg_lat = sum(p['lat'] for p in coord_pairs) / len(coord_pairs)
                        avg_lng = sum(p['lng'] for p in coord_pairs) / len(coord_pairs)
                        location_data['latitude'] = avg_lat
                        location_data['longitude'] = avg_lng
                        location_data['coordinate_count'] = len(coord_pairs)
                    
                    locations.append(location_data)
        
        if not locations:
            # Log the raw content for debugging (first 1000 chars)
            debug_content = content_str[:1000] if len(content_str) > 1000 else content_str
            raise HTTPException(
                status_code=400, 
                detail=f"Nenhuma localização válida encontrada no arquivo KML. Verifique se o arquivo contém elementos Placemark com coordenadas válidas. Debug: {debug_content}"
            )
        
        # Save to database
        kml_data = {
            "id": str(uuid.uuid4()),
            "filename": file.filename,
            "uploaded_by": admin_user.username,
            "uploaded_at": datetime.now(timezone.utc),
            "locations": locations,
            "total_locations": len(locations),
            "status": "active"
        }
        
        await db.kml_data.insert_one(kml_data)
        
        return {
            "message": f"Arquivo KML processado com sucesso! {len(locations)} localizações encontradas.",
            "kml_id": kml_data["id"],
            "total_locations": len(locations),
            "locations": locations[:10]  # Return first 10 as preview
        }
        
    except ET.ParseError as e:
        raise HTTPException(status_code=400, detail=f"Arquivo KML inválido ou corrompido: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo KML: {str(e)}")

@api_router.get("/kml/locations")
async def get_kml_locations(current_user: User = Depends(get_current_user)):
    kml_files = await db.kml_data.find({"status": "active"}).to_list(length=None)
    
    all_locations = []
    for kml_file in kml_files:
        for location in kml_file.get("locations", []):
            location["source_file"] = kml_file["filename"]
            location["uploaded_by"] = kml_file["uploaded_by"]
            all_locations.append(location)
    
    return all_locations

@api_router.delete("/admin/kml/{kml_id}")
async def delete_kml_data(kml_id: str, admin_user: User = Depends(get_admin_user)):
    result = await db.kml_data.delete_one({"id": kml_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dados KML não encontrados")
    
    return {"message": "Dados KML excluídos com sucesso"}

@api_router.get("/kml/search")
async def search_kml_locations(
    query: str,
    limit: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Search KML locations by name or description - only return results for specific searches"""
    if not query or len(query.strip()) < 2:
        raise HTTPException(status_code=400, detail="Query deve ter pelo menos 2 caracteres")
    
    query_lower = query.strip().lower()
    
    # Get all KML files
    kml_files = await db.kml_data.find({"status": "active"}).to_list(length=None)
    
    matching_locations = []
    for kml_file in kml_files:
        for location in kml_file.get("locations", []):
            # Search in name and description
            name_match = query_lower in location.get("name", "").lower()
            desc_match = query_lower in location.get("description", "").lower()
            
            if name_match or desc_match:
                location_data = {
                    "id": f"{kml_file['id']}_{len(matching_locations)}",
                    "name": location.get("name"),
                    "description": location.get("description"),
                    "latitude": location.get("latitude"),
                    "longitude": location.get("longitude"),
                    "source_file": kml_file["filename"],
                    "uploaded_by": kml_file["uploaded_by"]
                }
                matching_locations.append(location_data)
                
                if len(matching_locations) >= limit:
                    break
        
        if len(matching_locations) >= limit:
            break
    
    return {
        "query": query,
        "total_found": len(matching_locations),
        "locations": matching_locations
    }

@api_router.post("/kml/locations/{location_id}/observations")
async def add_location_observation(
    location_id: str,
    observation_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Add user observation to a specific location"""
    observation_text = observation_data.get("observation", "").strip()
    
    if not observation_text:
        raise HTTPException(status_code=400, detail="Observação não pode estar vazia")
    
    # Create observation record
    observation = {
        "id": str(uuid.uuid4()),
        "location_id": location_id,
        "user_id": current_user.id,
        "username": current_user.username,
        "observation": observation_text,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.location_observations.insert_one(observation)
    
    return {
        "message": "Observação adicionada com sucesso",
        "observation_id": observation["id"]
    }

@api_router.get("/kml/locations/{location_id}/observations")
async def get_location_observations(
    location_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get all observations for a specific location"""
    observations = await db.location_observations.find({
        "location_id": location_id
    }, {"_id": 0}).sort("created_at", -1).to_list(length=None)
    
    return observations

@api_router.delete("/kml/observations/{observation_id}")
async def delete_observation(
    observation_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete user's own observation"""
    observation = await db.location_observations.find_one({"id": observation_id})
    
    if not observation:
        raise HTTPException(status_code=404, detail="Observação não encontrada")
    
    # Users can only delete their own observations, admins can delete any
    if observation["user_id"] != current_user.id and current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Você só pode excluir suas próprias observações")
    
    await db.location_observations.delete_one({"id": observation_id})
    
    return {"message": "Observação excluída com sucesso"}

# Endpoints para perfil do usuário
@api_router.put("/user/change-password")
async def change_user_password(password_change: PasswordChange, current_user: User = Depends(get_current_user)):
    # Verificar senha atual
    if not verify_password(password_change.current_password, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    # Validar nova senha
    if len(password_change.new_password) < 4:
        raise HTTPException(status_code=400, detail="New password must be at least 4 characters")
    
    # Atualizar senha
    hashed_password = get_password_hash(password_change.new_password)
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {"hashed_password": hashed_password}}
    )
    
    return {"message": "Password changed successfully"}

@api_router.get("/reports/timeline")
async def get_timeline_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    from datetime import datetime, timezone, timedelta
    import calendar
    
    # Default to last 6 months if no dates provided
    if not end_date:
        end_date = datetime.now(timezone.utc)
    else:
        end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    
    if not start_date:
        start_date = end_date - timedelta(days=180)  # 6 months
    else:
        start_date = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
    
    # Group pendencies by month
    pipeline = [
        {
            "$match": {
                "created_at": {"$gte": start_date, "$lte": end_date}
            }
        },
        {
            "$group": {
                "_id": {
                    "year": {"$year": "$created_at"},
                    "month": {"$month": "$created_at"}
                },
                "total": {"$sum": 1},
                "pending": {
                    "$sum": {"$cond": [{"$eq": ["$status", "Pendente"]}, 1, 0]}
                },
                "finished": {
                    "$sum": {"$cond": [{"$eq": ["$status", "Finalizado"]}, 1, 0]}
                },
                "approved": {
                    "$sum": {"$cond": [{"$eq": ["$validation_status", "APPROVED"]}, 1, 0]}
                }
            }
        },
        {"$sort": {"_id.year": 1, "_id.month": 1}}
    ]
    
    results = await db.pendencias.aggregate(pipeline).to_list(100)
    
    # Format results
    timeline_data = []
    for result in results:
        month_name = calendar.month_name[result["_id"]["month"]]
        timeline_data.append({
            "period": f"{month_name} {result['_id']['year']}",
            "year": result["_id"]["year"],
            "month": result["_id"]["month"],
            "total": result["total"],
            "pending": result["pending"],
            "finished": result["finished"],
            "approved": result["approved"]
        })
    
    return timeline_data

@api_router.get("/reports/distribution")
async def get_distribution_report(current_user: User = Depends(get_current_user)):
    
    # Distribution by type
    type_pipeline = [
        {
            "$group": {
                "_id": "$tipo",
                "count": {"$sum": 1}
            }
        }
    ]
    
    # Distribution by site
    site_pipeline = [
        {
            "$group": {
                "_id": "$site",
                "count": {"$sum": 1}
            }
        },
        {"$sort": {"count": -1}},
        {"$limit": 10}  # Top 10 sites
    ]
    
    # Distribution by status
    status_pipeline = [
        {
            "$group": {
                "_id": "$status",
                "count": {"$sum": 1}
            }
        }
    ]
    
    type_results = await db.pendencias.aggregate(type_pipeline).to_list(100)
    site_results = await db.pendencias.aggregate(site_pipeline).to_list(100)
    status_results = await db.pendencias.aggregate(status_pipeline).to_list(100)
    
    return {
        "by_type": [{"type": r["_id"], "count": r["count"]} for r in type_results],
        "by_site": [{"site": r["_id"], "count": r["count"]} for r in site_results],
        "by_status": [{"status": r["_id"], "count": r["count"]} for r in status_results]
    }

@api_router.get("/reports/performance")
async def get_performance_report(current_user: User = Depends(get_current_user)):
    from datetime import datetime, timezone, timedelta
    
    # Last 30 days
    end_date = datetime.now(timezone.utc)
    start_date = end_date - timedelta(days=30)
    
    # Top performers (creators)
    creators_pipeline = [
        {
            "$match": {
                "created_at": {"$gte": start_date, "$lte": end_date}
            }
        },
        {
            "$group": {
                "_id": "$usuario_criacao",
                "created": {"$sum": 1},
                "approved": {
                    "$sum": {"$cond": [{"$eq": ["$validation_status", "APPROVED"]}, 1, 0]}
                }
            }
        },
        {"$sort": {"created": -1}},
        {"$limit": 10}
    ]
    
    # Top performers (finalizers)
    finalizers_pipeline = [
        {
            "$match": {
                "data_finalizacao": {"$gte": start_date, "$lte": end_date},
                "status": "Finalizado"
            }
        },
        {
            "$group": {
                "_id": "$usuario_finalizacao",
                "finished": {"$sum": 1},
                "approved": {
                    "$sum": {"$cond": [{"$eq": ["$validation_status", "APPROVED"]}, 1, 0]}
                }
            }
        },
        {"$sort": {"finished": -1}},
        {"$limit": 10}
    ]
    
    creators_results = await db.pendencias.aggregate(creators_pipeline).to_list(100)
    finalizers_results = await db.pendencias.aggregate(finalizers_pipeline).to_list(100)
    
    # Format results with approval rates
    top_creators = []
    for result in creators_results:
        approval_rate = (result["approved"] / result["created"] * 100) if result["created"] > 0 else 0
        top_creators.append({
            "username": result["_id"],
            "created": result["created"],
            "approved": result["approved"],
            "approval_rate": round(approval_rate, 1)
        })
    
    top_finalizers = []
    for result in finalizers_results:
        approval_rate = (result["approved"] / result["finished"] * 100) if result["finished"] > 0 else 0
        top_finalizers.append({
            "username": result["_id"],
            "finished": result["finished"],
            "approved": result["approved"],
            "approval_rate": round(approval_rate, 1)
        })
    
    return {
        "top_creators": top_creators,
        "top_finalizers": top_finalizers,
        "period": "Last 30 days"
    }

@api_router.get("/user/stats")
async def get_user_stats(current_user: User = Depends(get_current_user)):
    from datetime import datetime, timezone
    import calendar
    
    # Get current month/year
    now = datetime.now(timezone.utc)
    current_month = now.month
    current_year = now.year
    
    # Stats for current month
    start_date = datetime(current_year, current_month, 1, tzinfo=timezone.utc)
    if current_month == 12:
        end_date = datetime(current_year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        end_date = datetime(current_year, current_month + 1, 1, tzinfo=timezone.utc)
    
    # Pendências criadas pelo usuário no mês
    created_count = await db.pendencias.count_documents({
        "usuario_criacao": current_user.username,
        "created_at": {"$gte": start_date, "$lt": end_date}
    })
    
    # Pendências finalizadas pelo usuário no mês
    finished_count = await db.pendencias.count_documents({
        "usuario_finalizacao": current_user.username,
        "data_finalizacao": {"$gte": start_date, "$lt": end_date},
        "status": "Finalizado"
    })
    
    # Pendências aprovadas pelo admin que o usuário criou
    approved_created_count = await db.pendencias.count_documents({
        "usuario_criacao": current_user.username,
        "created_at": {"$gte": start_date, "$lt": end_date},
        "validation_status": "APPROVED"
    })
    
    # Pendências aprovadas pelo admin que o usuário finalizou
    approved_finished_count = await db.pendencias.count_documents({
        "usuario_finalizacao": current_user.username,
        "data_finalizacao": {"$gte": start_date, "$lt": end_date},
        "status": "Finalizado",
        "validation_status": "APPROVED"
    })
    
    return {
        "month": calendar.month_name[current_month],
        "year": current_year,
        "created_count": created_count,
        "finished_count": finished_count,
        "approved_created_count": approved_created_count,
        "approved_finished_count": approved_finished_count
    }

@api_router.get("/stats/monthly")
async def get_monthly_stats(current_user: User = Depends(get_current_user)):
    from datetime import datetime, timezone
    import calendar
    
    # Get current month/year
    now = datetime.now(timezone.utc)
    current_month = now.month
    current_year = now.year
    
    # Stats for current month
    start_date = datetime(current_year, current_month, 1, tzinfo=timezone.utc)
    if current_month == 12:
        end_date = datetime(current_year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        end_date = datetime(current_year, current_month + 1, 1, tzinfo=timezone.utc)
    
    # Most created pendencias this month (only validated by admin)
    created_pipeline = [
        {"$match": {
            "created_at": {"$gte": start_date, "$lt": end_date},
            "validation_status": "APPROVED"
        }},
        {"$group": {"_id": "$usuario_criacao", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 1}
    ]
    
    # Most finished pendencias this month (only validated by admin)
    finished_pipeline = [
        {"$match": {
            "data_finalizacao": {"$gte": start_date, "$lt": end_date},
            "status": "Finalizado",
            "validation_status": "APPROVED"
        }},
        {"$group": {"_id": "$usuario_finalizacao", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 1}
    ]
    
    most_created = list(await db.pendencias.aggregate(created_pipeline).to_list(1))
    most_finished = list(await db.pendencias.aggregate(finished_pipeline).to_list(1))
    
    return {
        "month": calendar.month_name[current_month],
        "year": current_year,
        "most_created": most_created[0] if most_created else None,
        "most_finished": most_finished[0] if most_finished else None
    }

@api_router.get("/pendencias/export")
async def export_pendencias(
    site: Optional[str] = None,
    tipo: Optional[str] = None,
    status: Optional[str] = None,
    admin_user: User = Depends(get_admin_user)
):
    # Build query
    query = {}
    if site:
        query["site"] = site
    if tipo:
        query["tipo"] = tipo
    if status:
        query["status"] = status
    
    # Get data
    pendencias = await db.pendencias.find(query).sort("created_at", -1).to_list(1000)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pendências"
    
    # Headers
    headers = [
        "ID", "Site", "Data/Hora", "Tipo", "Subtipo", 
        "Observações", "Status", "Usuário Criação", 
        "Usuário Finalização", "Data Finalização", "Informações Fechamento", "Foto Fechamento"
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for row, pendencia in enumerate(pendencias, 2):
        ws.cell(row=row, column=1, value=pendencia["id"])
        ws.cell(row=row, column=2, value=pendencia["site"])
        ws.cell(row=row, column=3, value=pendencia["data_hora"].strftime("%d/%m/%Y %H:%M") if pendencia.get("data_hora") else "")
        ws.cell(row=row, column=4, value=pendencia["tipo"])
        ws.cell(row=row, column=5, value=pendencia["subtipo"])
        ws.cell(row=row, column=6, value=pendencia["observacoes"])
        ws.cell(row=row, column=7, value=pendencia["status"])
        ws.cell(row=row, column=8, value=pendencia["usuario_criacao"])
        ws.cell(row=row, column=9, value=pendencia.get("usuario_finalizacao", ""))
        ws.cell(row=row, column=10, value=pendencia["data_finalizacao"].strftime("%d/%m/%Y %H:%M") if pendencia.get("data_finalizacao") else "")
        ws.cell(row=row, column=11, value=pendencia.get("informacoes_fechamento", ""))
        ws.cell(row=row, column=12, value="Sim" if pendencia.get("foto_fechamento_base64") else "Não")
    
    # Auto-adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to temporary file
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        return FileResponse(
            tmp.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="pendencias.xlsx"
        )


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()